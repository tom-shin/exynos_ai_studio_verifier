#!/usr/bin/env python3
import threading

DEBUG = False

import warnings

warnings.filterwarnings("ignore", category=FutureWarning)

if not DEBUG:
    import onnx
    import onnxruntime
    import onnxsim

import logging
import easygui
import platform
import time
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter
from ruamel.yaml import YAML
from io import StringIO
import numpy as np
import uuid
from typing import Tuple, List, Dict

from matplotlib.backends.backend_qt5agg import FigureCanvasQTAgg as FigureCanvas
from matplotlib.figure import Figure

from PyQt5.QtCore import pyqtSignal, QObject, QThread, QEventLoop
from PyQt5 import QtWidgets, QtCore, QtGui
from PyQt5.QtWidgets import QFileDialog, QApplication

from source.__init__ import *

from source.source_files.execute_verify import get_image_info_from_dockerImg, Load_Target_Dir_Thread, \
    start_docker_desktop, set_model_config

from source.source_files.run_enntest import upgrade_local_run_enntest, upgrade_remote_run_enntest

if getattr(sys, 'frozen', False):
    BASE_DIR = sys._MEIPASS
else:
    BASE_DIR = os.path.dirname(os.path.abspath(__file__))  # 일반 Python 스크립트 실행 시


def load_module_func(module_name):
    mod = __import__(f"{module_name}", fromlist=[module_name])
    return mod


class WorkerThread(QThread):
    finished = pyqtSignal(bool, list, list)  # ret, failed_pairs, memory_profile 전달

    def __init__(self, remote, nnc_files, input_golden_pairs, current_binary_pos, out_dir, enntest_combo, deviceID):
        super().__init__()
        self.remote = remote
        self.nnc_files = nnc_files
        self.input_golden_pairs = input_golden_pairs
        self.current_binary_pos = current_binary_pos
        self.out_dir = out_dir
        self.enntest_combo = enntest_combo
        self.deviceID = deviceID
        self.enntest_profile_iter = [10000, 20000, 30000]

    @staticmethod
    def check_retry(memory_profile):
        try:
            start_idx = memory_profile.index("Start")
            end_idx = memory_profile.index("End")
        except ValueError:
            start_idx = 0
            end_idx = len(memory_profile)

        values_between = memory_profile[start_idx + 1:end_idx]

        # 값이 없거나 모든 값이 '0'인지 확인
        if not values_between or all(value == '0' for value in values_between):
            return True
        return False

    def run(self):
        ret = False
        failed_pairs = []
        memory_profile = []

        # 추론 시간이 너무 짧아 메모리 사용량 체크를 못할 경우 iteration 증가해서 재 평가
        for profile_iter in self.enntest_profile_iter:
            if self.remote:
                ret, failed_pairs, memory_profile = upgrade_remote_run_enntest(
                    nnc_files=self.nnc_files, input_golden_pairs=self.input_golden_pairs, current_binary_pos=self.current_binary_pos,
                    out_dir=self.out_dir, profile_iter=profile_iter, deviceID=self.deviceID
                )

            else:
                ret, failed_pairs, memory_profile = upgrade_local_run_enntest(
                    nnc_files=self.nnc_files, input_golden_pairs=self.input_golden_pairs,
                    current_binary_pos=self.current_binary_pos,
                    out_dir=self.out_dir, profile_iter=profile_iter, deviceID=self.deviceID
                )
            if not self.check_retry(memory_profile=memory_profile):
                break

        self.finished.emit(ret, failed_pairs, memory_profile)


class EmittingStream(QObject):
    textWritten = pyqtSignal(str)

    def write(self, text):
        self.textWritten.emit(str(text))

    def flush(self):
        pass


class ImageLoadThread(QThread):
    send_finish_ui_sig = QtCore.pyqtSignal(bool, str)

    def __init__(self, parent, arg1, arg2):
        super().__init__(parent)
        self.parent = parent
        self.command = arg1
        self.run_time = arg2

    def run(self):
        start_time = time.time()
        load_result, _, _ = user_subprocess(cmd=self.command, run_time=self.run_time)
        elapsed_time = time.time() - start_time

        days = elapsed_time // (24 * 3600)
        remaining_secs = elapsed_time % (24 * 3600)
        hours = remaining_secs // 3600
        remaining_secs %= 3600
        minutes = remaining_secs // 60
        seconds = remaining_secs % 60

        total_time = f"{int(days)}day {int(hours)}h {int(minutes)}m {int(seconds)}s"

        if any("Loaded image:" in line for line in load_result):
            self.send_finish_ui_sig.emit(True, total_time)
        else:
            self.send_finish_ui_sig.emit(False, "")

    def stop(self):
        self.quit()
        self.wait(3000)


class Model_Analyze_Thread(QThread):
    output_signal = pyqtSignal(str, tuple, int, str)
    output_signal_2 = pyqtSignal(tuple, str, dict)

    timeout_output_signal = pyqtSignal(str, tuple)
    finish_output_signal = pyqtSignal(bool)

    send_max_progress_cnt = pyqtSignal(int)
    send_set_text_signal = pyqtSignal(str)
    send_onnx_opset_ver_signal = pyqtSignal(tuple, dict)

    def __init__(self, parent=None, grand_parent=None, ssh_client=None, repo_tag=None):
        super().__init__()
        self.worker = None
        self.parent = parent
        self.grand_parent = grand_parent
        self._running = True
        self.ssh_client = ssh_client
        self.DockerImg = repo_tag
        self.timeout_expired = float(grand_parent.timeoutlineEdit.text().strip())

    @staticmethod
    def test_model_integrity(model_path: str) -> bool:
        try:
            # ONNX Runtime 세션 생성
            session = onnxruntime.InferenceSession(model_path)

            # 입력 텐서 정보 가져오기
            input_tensors = session.get_inputs()

            # 입력 데이터 생성 (zeros로 초기화)
            input_data = []
            for tensor in input_tensors:
                shape = tuple(map(int, tensor.shape))  # 문자열을 정수로 변환
                data = np.zeros(shape, dtype=np.float32)
                input_data.append(data)

            # 推论 수행
            outputs = session.run(None, dict(zip([t.name for t in input_tensors], input_data)))

            return True

        except Exception as e:
            PRINT_(f"에러 발생: {str(e)}")
            return False

    @staticmethod
    def get_opset_version(p_model, extension):
        opset_version = ""
        opset_domain = ""

        if extension.lower() != ".onnx":
            return opset_domain, opset_version

        model = onnx.load(p_model)
        opset_imports = model.opset_import

        for opset in opset_imports:
            domain = f"Domain: {opset.domain or 'ai.onnx'}\n"  # Operator set version
            opset_domain += domain

            ver = f"opset ver.: {opset.version}\n"  # Operator set version
            opset_version += ver

        return opset_domain, opset_version

    @staticmethod
    def check_model_precision(p_model, extension):
        quantization_type = ""

        if extension.lower() != ".onnx":
            return quantization_type

        model = onnx.load(p_model)

        # 1. 초기화 파라미터(weights) 체크
        initializer_types = set()
        for initializer in model.graph.initializer:
            initializer_types.add(initializer.data_type)

        # 2. 입력/출력 텐서 체크
        tensor_types = set()
        for value_info in model.graph.value_info:
            tensor_types.add(value_info.type.tensor_type.elem_type)

        # 3. 모델 입력/출력 체크
        for input_ in model.graph.input:
            tensor_types.add(input_.type.tensor_type.elem_type)
        for output in model.graph.output:
            tensor_types.add(output.type.tensor_type.elem_type)

        # ONNX의 데이터 타입 매핑
        onnx_dtype_map = {
            1: "FLOAT",
            2: "UINT8",
            3: "INT8",
            4: "UINT16",
            5: "INT16",
            6: "INT32",
            7: "INT64",
            10: "FLOAT16",
            11: "DOUBLE",
            12: "UINT32",
            13: "UINT64",
        }

        # ONNX의 데이터 타입 (1 = FLOAT)
        is_fp32 = all(dtype == 1 for dtype in initializer_types) and \
                  all(dtype == 1 for dtype in tensor_types)

        # 모델 연산자 수집
        model_ops = set(node.op_type for node in model.graph.node)

        def get_quantization_type():
            # 연산자 및 데이터 타입 기반 체크
            quant_signatures = {
                "INT8": {
                    "ops": ["QLinearConv", "QuantizeLinear", "DequantizeLinear"],
                    "dtype": 3  # INT8
                },
                "UINT8": {
                    "ops": ["QLinearConv", "QuantizeLinear", "DequantizeLinear"],
                    "dtype": 2  # UINT8
                },
                "INT4": {
                    "ops": ["DequantizeLinear", "QuantizeLinear"],
                    "dtype": 3,  # INT8 but with 4-bit quantization
                    "extra_check": lambda node: any("nbits" in attr.name and attr.i == 4
                                                    for attr in node.attribute)
                },
                "FP16": {
                    "ops": [],  # FP16 모델은 특별한 양자화 연산자 없음
                    "dtype": 10  # FLOAT16
                },
                "FP8": {
                    "ops": ["FP8Conv", "FP8MatMul"],  # FP8 특화 연산자
                    "dtype": None
                }
            }

            def check_weight_bits():
                # weight의 비트 수 확인
                for node in model.graph.node:
                    if hasattr(node, 'attribute'):
                        for attr in node.attribute:
                            if attr.name == 'weight_bits':
                                return attr.i
                return None

            # 양자화 타입 판별
            found_types = set()

            # 데이터 타입 기반 체크
            for dtype in initializer_types.union(tensor_types):
                if dtype == 10:  # FLOAT16
                    found_types.add("FP16")
                elif dtype == 3:  # INT8
                    bits = check_weight_bits()
                    if bits == 4:
                        found_types.add("INT4")
                    else:
                        found_types.add("INT8")
                elif dtype == 2:  # UINT8
                    found_types.add("UINT8")

            # 연산자 기반 체크
            if "QLinearConv" in model_ops:
                for node in model.graph.node:
                    if node.op_type == "QLinearConv":
                        # 입력 텐서의 데이터 타입으로 판별
                        for init in model.graph.initializer:
                            if init.name in node.input:
                                if init.data_type == 3:  # INT8
                                    found_types.add("INT8")
                                elif init.data_type == 2:  # UINT8
                                    found_types.add("UINT8")

            # FP8 특화 연산자 체크
            if any(op.startswith("FP8") for op in model_ops):
                found_types.add("FP8")

            if not found_types:
                return None
            elif len(found_types) == 1:
                quant_type = found_types.pop()
                method = "QLinearConv-based" if "QLinearConv" in model_ops else \
                    "QDQ-based" if "QuantizeLinear" in model_ops else \
                        "Native"
                return f"{quant_type} ({method} quantization)"
            else:
                return f"Mixed quantization types: {', '.join(sorted(found_types))}"

        quant_type = get_quantization_type()

        if is_fp32 and not quant_type:
            quantization_type = "Confirmed FP32 model (all tensors are FP32)"
        elif quant_type:
            quantization_type = quant_type
            # PRINT_(f"  • Model Type: {quant_type}")
            # 추가 정보 출력
            additional_msg = ", ".join(op for op in model_ops if op.startswith("Q") or "FP8" in op)
            quantization_type += f"\n- Quantization operators found:  {additional_msg}"
            # PRINT_("    - Quantization operators found:",
            #       ", ".join(op for op in model_ops if op.startswith("Q") or "FP8" in op))
        else:
            additional_msg = ", ".join(onnx_dtype_map.get(t, f"Unknown({t})") for t in initializer_types)
            additional_msg2 = ", ".join(onnx_dtype_map.get(t, f"Unknown({t})") for t in tensor_types)
            quantization_type = f"Model Type: Mixed precision or unknown\n,  - Found initializer types: {additional_msg}\n,  - Found tensor types: {additional_msg2}"

            # PRINT_("  • Model Type: Mixed precision or unknown")
            # PRINT_("    - Found initializer types:",
            #       ", ".join(onnx_dtype_map.get(t, f"Unknown({t})") for t in initializer_types))
            # PRINT_("    - Found tensor types:",
            #       ", ".join(onnx_dtype_map.get(t, f"Unknown({t})") for t in tensor_types))

        return quantization_type

    @staticmethod
    def check_model_validation_f(p_model, extension):
        model_validation = ""
        if extension.lower() != ".onnx":
            return model_validation

        # ONNX 모델이 구조적으로 ONNX 스팩을 만족하는지 점검
        model = onnx.load(p_model)

        try:
            onnx.checker.check_model(model)
            model_validation = "Passed"

        except onnx.checker.ValidationError as e:
            model_validation = f"[Error] {str(e)}"

        except Exception as e:
            model_validation = f"[Error] {str(e)}"

        return model_validation

    @staticmethod
    def test_model_integrity(p_model, extension):
        is_integrity_ok = ""
        onnx_model_runtime_inference_check_log = ""
        if extension.lower() != ".onnx":
            return is_integrity_ok, onnx_model_runtime_inference_check_log

        try:
            # ONNX Runtime 세션 생성
            session = onnxruntime.InferenceSession(p_model)

            # 입력 텐서 정보 가져오기
            input_tensors = session.get_inputs()

            # 입력 데이터 생성 (zeros로 초기화)
            input_data = []
            for tensor in input_tensors:
                shape = tuple(map(int, tensor.shape))  # 문자열을 정수로 변환
                data = np.zeros(shape, dtype=np.float32)
                input_data.append(data)

            # 推论 수행
            outputs = session.run(None, dict(zip([t.name for t in input_tensors], input_data)))
            is_integrity_ok = "Passed"

        except Exception as e:
            PRINT_(f"[Error]: {e}")
            is_integrity_ok = "Failed"
            onnx_model_runtime_inference_check_log = str(e)

        return is_integrity_ok, onnx_model_runtime_inference_check_log

    @staticmethod
    def analyze_tensor_shapes(tensor) -> Dict[str, List]:
        """텐서의 shape 정보를 분석합니다."""
        shape_info = []
        is_dynamic = False

        for dim in tensor.type.tensor_type.shape.dim:
            if dim.dim_param:  # 심볼릭 차원
                shape_info.append(f"dynamic({dim.dim_param})")
                is_dynamic = True
            elif dim.dim_value:  # 고정 차원
                shape_info.append(str(dim.dim_value))
            else:  # 미정의 차원
                shape_info.append("dynamic")
                is_dynamic = True

        return {
            "name": tensor.name,
            "shape": shape_info,
            "is_dynamic": is_dynamic
        }

    def check_model_directly(self, model_path: str) -> Tuple[bool, List[Dict]]:
        """모델을 직접 분석하여 동적 shape를 검사합니다."""
        model = onnx.load(model_path)

        # try:
        #     onnx.checker.check_model(model)
        #     PRINT_("  • Model check passed ✓")
        # except onnx.checker.ValidationError as e:
        #     PRINT_(f"  • Validation Error: {str(e)}")
        # except Exception as e:
        #     PRINT_(f"  • Error: {str(e)}")

        has_dynamic = False
        tensor_info = []

        # 입력 텐서 분석
        for input in model.graph.input:
            info = self.analyze_tensor_shapes(input)
            info["type"] = "input"
            tensor_info.append(info)
            if info["is_dynamic"]:
                has_dynamic = True

        # 출력 텐서 분석
        for output in model.graph.output:
            info = self.analyze_tensor_shapes(output)
            info["type"] = "output"
            tensor_info.append(info)
            if info["is_dynamic"]:
                has_dynamic = True

        return has_dynamic, tensor_info

    @staticmethod
    def run_simplifier(model_path: str) -> List[str]:
        """ONNX Simplifier를 실행하고 발생하는 메시지를 수집합니다."""
        try:
            # simplified 폴더 생성 (현재 작업 디렉토리 기준)
            output_dir = os.path.join(os.getcwd(), "Result", "simplified")
            if not os.path.exists(output_dir):
                os.makedirs(output_dir)
                messages = [f"Created 'simplified' Directory: {output_dir}"]
            else:
                messages = []

            # 출력 파일 경로 생성
            base_name = os.path.splitext(os.path.basename(model_path))[0]
            output_path = os.path.join(output_dir, f"{base_name}_simplified.onnx")
            # messages.append(f"단순화된 모델이 다음 경로에 저장될 예정: {output_path}")

            model = onnx.load(model_path)

            try:
                simplified_model, check = onnxsim.simplify(
                    model,
                    skipped_optimizers=['eliminate_identity']
                )

                # 단순화된 모델 저장
                onnx.save(simplified_model, output_path)
                messages.append(f"Executed Simplifier: (Return Value: {check})")
                messages.append(f"Saved Path: {output_path}")

            except Exception as e:
                messages.append(f"Exception Error: {str(e)}")

            return messages

        except Exception as e:
            return [f"Exception Error: {str(e)}"]

    @staticmethod
    def remove_all_container():
        cmd = "docker rm -f $(docker ps -aq)"
        user_subprocess(cmd=f"{cmd}", run_time=False, log=False)
        # for container_name in self.container_trace:
        #     user_subprocess(cmd=f"docker stop {container_name}", run_time=False, log=False)
        #     user_subprocess(cmd=f"docker rm {container_name}", run_time=False, log=False)

    def create_container(self, target_widget=None):
        cwd, _ = separate_folders_and_files(target_widget[0].pathlineEdit.text())

        # 현재 경로를 WSL 경로로 변환
        drive, path = os.path.splitdrive(os.path.dirname(cwd))

        env = check_environment()
        if env == "Windows":
            self.Shared_Volume = "/mnt/" + drive.lower().replace(":", "") + path.replace("\\", "/")
        else:
            self.Shared_Volume = path.replace("\\", "/")

        ContainerName = f"container_{uuid.uuid4().hex}"  # 고유값 필요요

        base_cmd = [
            "docker",
            "run",
            "-d",
            "--name",
            ContainerName,
            "--security-opt", "seccomp:unconfined",  # 보안 프로필 해제 (2번 항목)
            "--cap-add=ALL",  # 모든 리눅스 기능 추가 (2번 항목)
            "--privileged",  # 권한 강화 (2번 항목)
            "--net", "host",  # 호스트 네트워크 사용 (3번 항목)
            "--ipc", "host",  # 호스트 IPC 네임스페이스 사용 (3번 항목)           
            "-v", f"{self.Shared_Volume}:/workspace",  # 볼륨 마운트 (이미 포함됨)
            "-v", "/etc/timezone:/etc/timezone",  # 타임존 설정 (6번 항목)
            "-w", "/workspace",  # 작업 디렉터리 설정 (6번 항목)           
            self.DockerImg,
            "/bin/bash",
            "-c",
            "tail -f /dev/null"
        ]

        if self.grand_parent.cpuradioButton.isChecked():  # CPU
            # For CPU
            cmd = base_cmd[:3] + base_cmd[3:]  # No changes needed, just reuse the base command
        else:
            # For GPU
            cmd = base_cmd[:3] + ["--gpus", "all"] + base_cmd[3:]  # Add GPU specific option

        user_subprocess(cmd=cmd, run_time=False, log=False, shell=False)

        return ContainerName, cwd

    def get_thread_status(self):
        return self._running

    @staticmethod
    def get_model_execute(model=None):
        return model[0].scenario_checkBox.isChecked()

    def get_checked_model_lists(self):
        CommandLists = []
        for i in range(0, 7):
            check_box_name = f"cmd{i}"  # 체크박스의 이름을 동적으로 생성
            check_box = getattr(self.grand_parent, check_box_name, None)  # 객체 가져오기
            if check_box and check_box.isChecked():  # 체크박스가 존재하고 선택되었는지 확인
                CommandLists.append(check_box.text())  # 체크박스의 텍스트를 리스트에 추가
        # PRINT_("++++++++++++++\n", CommandLists)
        return CommandLists

    def get_basic_onnx_model_information(self, model=None, extension=None):
        if extension.lower() != ".onnx":
            return None

        onnx_info = {}

        # 0. Model Check 분석
        onnx_validation = self.check_model_validation_f(p_model=model, extension=extension)  # check.model
        onnx_info["model check"] = [onnx_validation, ""]

        # 1. Direct Shape 분석
        is_dynamic = "Static shape"
        has_dynamic_direct, tensor_info = self.check_model_directly(model_path=model)
        if has_dynamic_direct:
            is_dynamic = "Dynamic shape"
        onnx_info["direct shape"] = [is_dynamic, tensor_info]

        # 2. Simplifier 실행
        sim_msg = ""
        simplifier_messages = self.run_simplifier(model_path=model)
        for msg in simplifier_messages:
            sim_msg += f"{msg}\n"
        onnx_info["simplifier"] = [sim_msg, ""]

        # 3. 모델 무결성 테스트
        is_integrity_ok, is_integrity_ok_log = self.test_model_integrity(p_model=model, extension=extension)
        onnx_info["integrity"] = [is_integrity_ok, is_integrity_ok_log]

        # 4. 모델 정밀도 분석 추가
        onnx_model_type = self.check_model_precision(p_model=model, extension=extension)
        onnx_info["precision"] = [onnx_model_type, ""]

        # 5. 모델 domain, opset 버전 출력
        model_domain, opset_ver = self.get_opset_version(p_model=model, extension=extension)
        onnx_info["domain"] = [model_domain, ""]
        onnx_info["opset_version"] = [opset_ver, ""]

        return onnx_info

    def check_enntools_init(self, init_log_path=None, model=None, filename=None):
        result = 'Fail'
        parameter_set = OrderedDict()

        check_DATA_dir = os.path.join(init_log_path, "DATA").replace("\\", "/")
        target_config = os.path.join(init_log_path, f"{filename}.yaml").replace("\\", "/")

        if os.path.isdir(check_DATA_dir) and os.path.isfile(target_config):
            result = "Success"

        if self.grand_parent.modifiedradioButton.isChecked():
            repo_tag = self.grand_parent.dockerimagecomboBox.currentText()
            tag = int(repo_tag.split(":")[1].split(".")[0])

            if tag >= 7:
                src_config = os.path.join(BASE_DIR, "model_configuration",
                                          "Ver2.0_model_config_new.yaml").replace("\\", "/")
            else:
                src_config = os.path.join(BASE_DIR, "model_configuration",
                                          "Ver1.0_model_config_new.yaml").replace("\\", "/")

            parameter_set = set_model_config(grand_parent=self.grand_parent, my_src_config=src_config,
                                             target_config=target_config,
                                             model_name=model)
        return result, parameter_set

    @staticmethod
    def check_enntools_conversion(cwd=None):
        result = 'Fail'
        error_contents_dict = ''
        check_log = os.path.join(cwd, "Converter_result", ".log")
        error_keywords = keyword["error_keyword"]

        if os.path.exists(check_log):
            ret, error_contents_dict = upgrade_check_for_specific_string_in_files(check_log,
                                                                                  check_keywords=error_keywords)
            if len(ret) == 0:
                result = "Success"

        return result, error_contents_dict

    @staticmethod
    def check_enntools_compile(cwd=None):
        result = 'Fail'
        error_contents_dict = ''
        check_log = os.path.join(cwd, "Compiler_result", ".log")
        error_keywords = keyword["error_keyword"]

        if os.path.exists(check_log):
            ret, error_contents_dict = upgrade_check_for_specific_string_in_files(check_log,
                                                                                  check_keywords=error_keywords)
            if len(ret) == 0:
                result = "Success"

        return result, error_contents_dict

    @staticmethod
    def check_enntools_cmd_result(check_log=''):
        result = 'Fail'
        error_contents_dict = {}
        error_keywords = keyword["error_keyword"]

        if os.path.exists(check_log):
            ret, error_contents_dict = upgrade_check_for_specific_string_in_files(check_log,
                                                                                  check_keywords=error_keywords)
            if len(ret) == 0:
                result = "Success"
        else:
            result = "Result directory not existed"

        return result, error_contents_dict

    def run_task(self, nnc_files, input_golden_pairs, current_binary_pos, out_dir, enntest_combo, deviceID):
        remote = True
        if not self.grand_parent.remoteradioButton.isChecked():
            remote = False

        self.worker = WorkerThread(remote, nnc_files, input_golden_pairs, current_binary_pos, out_dir, enntest_combo,
                                   deviceID)
        loop = QEventLoop()  # 이벤트 루프 생성

        result = {}

        def handle_finished(ret, failed_pairs, memory_profile):
            result['ret'] = ret
            result['failed_pairs'] = failed_pairs
            result['memory_profile'] = memory_profile
            loop.quit()

        self.worker.finished.connect(handle_finished)

        self.worker.start()
        loop.exec_()

        return result['ret'], result['failed_pairs'], result['memory_profile']

        # if self.grand_parent.remoteradioButton.isChecked():
        #     ret, failed_pairs, memory_profile = upgrade_remote_run_enntest(nnc_files, input_golden_pairs,
        #                                                                    current_binary_pos,
        #                                                                    out_dir,
        #                                                                    enntest_combo,
        #                                                                    deviceID)
        # else:
        #     ret, failed_pairs, memory_profile = upgrade_local_run_enntest(nnc_files, input_golden_pairs,
        #                                                                   current_binary_pos,
        #                                                                   out_dir,
        #                                                                   enntest_combo,
        #                                                                   deviceID)
        # return ret, failed_pairs, memory_profile

    def upgrade_execute_enntest_ondevice(self, TestResult=None, cwd=None):
        memory_profile = []
        failed_pairs = []
        nnc_model_path = os.path.join(cwd, "Compiler_result").replace("\\", "/")
        nnc_files = []

        if TestResult["enntools compile"][0] == "Success":
            for file in os.listdir(nnc_model_path):
                if file.endswith('.nnc') and os.path.isfile(os.path.join(nnc_model_path, file)):
                    filename = os.path.join(nnc_model_path, file).replace("\\", "/")
                    nnc_files.append(filename)

            input_golden_path = os.path.join(cwd, "Converter_result").replace("\\", "/")
            input_golden_pairs, current_binary_pos = upgrade_find_paired_files(input_golden_path)

            out_dir = os.path.join(cwd, "Enntester_result").replace("\\", "/")
            CheckDir(out_dir)

            if len(nnc_files) != 0 and len(input_golden_pairs) != 0:
                if self.grand_parent.remoteradioButton.isChecked():
                    deviceID = self.grand_parent.sshdevicelineEdit.text().strip()
                else:
                    deviceID = self.grand_parent.localdeviceidlineEdit.text().strip()

                ret, failed_pairs, memory_profile = self.run_task(nnc_files=nnc_files,
                                                                  input_golden_pairs=input_golden_pairs,
                                                                  current_binary_pos=current_binary_pos,
                                                                  out_dir=out_dir,
                                                                  enntest_combo=self.grand_parent.enntestcomboBox.currentText(),
                                                                  deviceID=deviceID)

                # if self.grand_parent.remoteradioButton.isChecked():
                #     ret, failed_pairs, memory_profile = upgrade_remote_run_enntest(nnc_files, input_golden_pairs,
                #                                                                    current_binary_pos,
                #                                                                    out_dir,
                #                                                                    self.grand_parent.enntestcomboBox.currentText(),
                #                                                                    deviceID=self.grand_parent.sshdevicelineEdit.text().strip())
                # else:
                #     ret, failed_pairs, memory_profile = upgrade_local_run_enntest(nnc_files, input_golden_pairs,
                #                                                                   current_binary_pos,
                #                                                                   out_dir,
                #                                                                   self.grand_parent.enntestcomboBox.currentText(),
                #                                                                   deviceID=self.grand_parent.localdeviceidlineEdit.text().strip())

                if ret:
                    result = "Success"
                else:
                    result = "Fail"
                    if len(failed_pairs) == 0:
                        result = "Server Connection Error"  # remote server failed
                    elif len(failed_pairs) == 1 and failed_pairs[0] == "check_device":
                        result = "Device Not Connected"  # remote server failed

            else:
                result = "[Warning !] Check Manually"
        else:
            result = "Skip"

        return result, failed_pairs, memory_profile

    def model_ai_studio_test(self, CommandLists=[], target_widget=None, executed_cnt=0, max_cnt=0):
        memory_profile = []

        def convert_changed_items_to_text(changed_items):
            # changed_items OrderedDict를 텍스트 형식으로 변환
            result_text = ""
            for section, keys in changed_items.items():  # OrderedDict의 섹션 순회
                result_text += f"Section: {section}\n"  # 섹션 이름 추가
                for key, values in keys.items():  # 각 섹션 내 키 순회
                    old_value = values['old_value']
                    new_value = values['new_value']
                    result_text += f"  Key: {key}\n"
                    result_text += f"    Old Value: {old_value}\n"
                    result_text += f"    New Value: {new_value}\n\n"

            return result_text

        def dict2string(err_log=None):
            text_to_display = ""
            for file, contexts in err_log.items():
                text_to_display += f"File: {file}\n"
                for context in contexts:
                    text_to_display += f"{context}\n{'-' * 40}\n"

            return text_to_display

        def list2string(failed_pair=None):
            string_ = ""
            for _in_bin, _golden_bin in failed_pair:
                string_ += f"[{os.path.basename(_in_bin)}, {os.path.basename(_golden_bin)}]\n"

            return string_

        TestResult = {
            # "enntools init": ["", ""],
            # "enntools conversion": ["", ""],
            # "enntools compile":["", ""],
            # "enntools estimation": ["", ""],
            # "enntools analysis": ["", ""],
            # "enntools profiling": ["", ""],
            # "enntest_ondevice": ["", ""]
        }

        ContainerName, cwd = self.create_container(target_widget=target_widget)

        start_T = time.time()
        for enntools_cmd in CommandLists:
            message = rf"{os.path.basename(cwd)} Executing {enntools_cmd} ...({executed_cnt}/{max_cnt})"
            self.send_set_text_signal.emit(message)

            if "enntest" in enntools_cmd:
                result, error_pair, memory_profile = self.upgrade_execute_enntest_ondevice(TestResult=TestResult,
                                                                                           cwd=cwd)

                if result == "Success":
                    log = "".join(error_pair)
                else:
                    PRINT_("+++++++++++++++++++++++++++++", error_pair)
                    if len(error_pair) != 0:
                        log = error_pair[-1]
                    else:
                        log = ""

                TestResult[enntools_cmd] = [result, log, memory_profile]

            else:

                CMD = [
                    "docker",
                    "exec",
                    "-i",
                    ContainerName,
                    "/bin/bash",
                    "-c",
                    # "ls -l /workspace"
                    f"cd /workspace/{os.path.basename(cwd)} && {enntools_cmd}"
                ]

                out, error, timeout_expired = user_subprocess(cmd=CMD, run_time=False, timeout=self.timeout_expired,
                                                              shell=False, log=True)

                env = check_environment()
                if env == "Linux":
                    post_cmd = [
                        "sudo",
                        "chmod",
                        "-R",
                        "777",
                        f"{self.Shared_Volume}"
                    ]
                    _, _, _ = user_subprocess(cmd=post_cmd, shell=False, timeout=self.timeout_expired, log=True)

                if timeout_expired:
                    self.timeout_output_signal.emit(enntools_cmd, target_widget)
                    break

                if "init" in enntools_cmd:
                    _, _model_ = separate_folders_and_files(target_widget[0].pathlineEdit.text())
                    _filename_, extension = separate_filename_and_extension(_model_)

                    result, parameter_set = self.check_enntools_init(init_log_path=cwd, model=_model_,
                                                                     filename=_filename_)
                    TestResult[enntools_cmd] = [result, convert_changed_items_to_text(parameter_set), memory_profile]
                else:
                    check_log = None
                    compile_profile_log = None
                    VisualProfiler_Summary_log = None

                    if "conversion" in enntools_cmd:
                        check_log = os.path.join(cwd, "Converter_result", ".log")

                    elif "compile" in enntools_cmd:
                        check_log = os.path.join(cwd, "Compiler_result", ".log")
                        compile_profile_log = os.path.join(cwd, "Compiler_result", "profile_log.txt")

                    elif "estimation" in enntools_cmd:
                        check_log = os.path.join(cwd, "Estimator_result", ".log")

                    elif "analysis" in enntools_cmd:
                        check_log = os.path.join(cwd, "Analyzer_result", ".log")

                    elif "profiling" in enntools_cmd:
                        check_log = os.path.join(cwd, "Profiler_result", ".log")
                        VisualProfiler_Summary_log = os.path.join(cwd, "Profiler_result", "VisualProfiler",
                                                                  "ResultProcess", "VisualProfiler_Summary.json")

                    if check_log is not None:
                        # PRINT_(check_log)
                        result, dict_log = self.check_enntools_cmd_result(check_log=check_log)
                        log = dict2string(dict_log)

                        if compile_profile_log is not None and os.path.exists(compile_profile_log):
                            log += f"\n\n[Profile Log]\n"

                            with open(compile_profile_log, 'r') as file:
                                lines = file.readlines()
                            for line in lines:
                                if "Unsupported" in line:  # "Unsupported" 키워드 포함 여부 확인
                                    log += line.strip() + "\n"  # 라인을 추가하고 줄 바꿈 추가

                        elif VisualProfiler_Summary_log is not None and os.path.exists(VisualProfiler_Summary_log):
                            log = ""

                            with open(VisualProfiler_Summary_log, 'r') as file:
                                lines = file.readlines()
                            for line in lines:
                                if "{" in line or "}" in line:
                                    continue
                                log += line.strip() + "\n"  # 라인을 추가하고 줄 바꿈 추가

                        TestResult[enntools_cmd] = [result, log, memory_profile]

            #  enntools cmd 실행될 때마다 update
            self.output_signal_2.emit(target_widget, enntools_cmd, TestResult)

        elapsed_time = self.convert_elapsedTime(start=start_T, finish=time.time())

        return elapsed_time, cwd

    def core_ai_studio_test(self, target_widget=None, CommandLists=[], executed_cnt=0, max_cnt=0):
        model = target_widget[0].pathlineEdit.text()
        name, ext = os.path.splitext(model)

        # 최재훈님 전달한 파일
        if not DEBUG:
            onnx_info = self.get_basic_onnx_model_information(model=model, extension=ext)
            if onnx_info is not None:
                self.send_onnx_opset_ver_signal.emit(target_widget, onnx_info)

        # enntools cmd test
        elapsed_time, cwd = self.model_ai_studio_test(CommandLists=CommandLists, target_widget=target_widget,
                                                      executed_cnt=executed_cnt, max_cnt=max_cnt)
        self.output_signal.emit(elapsed_time, target_widget, executed_cnt, cwd)

    def run(self):
        max_cnt = sum(1 for widget in self.parent.added_scenario_widgets if widget[0].scenario_checkBox.isChecked())
        self.send_max_progress_cnt.emit(max_cnt)

        executed_cnt = 0
        CommandLists = self.get_checked_model_lists()
        self.remove_all_container()

        for target_widget in self.parent.added_scenario_widgets:
            if not self.get_thread_status():
                break

            if not self.get_model_execute(model=target_widget):
                continue

            executed_cnt += 1
            self.remove_all_container()
            self.core_ai_studio_test(target_widget=target_widget, CommandLists=CommandLists, executed_cnt=executed_cnt,
                                     max_cnt=max_cnt)
            self.remove_all_container()

        self.finish_output_signal.emit(self._running)

    @staticmethod
    def convert_elapsedTime(start, finish):
        elapsed_time = finish - start

        days = elapsed_time // (24 * 3600)
        remaining_secs = elapsed_time % (24 * 3600)
        hours = remaining_secs // 3600
        remaining_secs %= 3600
        minutes = remaining_secs // 60
        seconds = remaining_secs % 60
        total_time = f"{int(days)}day {int(hours)}h {int(minutes)}m {int(seconds)}s"

        return total_time

    def re_config_yaml_file(self, target_widget):
        device = self.grand_parent.devicecomboBox.currentText()
        test_vector_gen = self.grand_parent.vectorcomboBox.currentText()
        debug = self.grand_parent.debugcomboBox.currentText()

        # 파일 경로 설정
        target_dir = target_widget[0].pathlineEdit
        filename = os.path.basename(target_dir)
        target_file = os.path.join(target_dir, filename).replace("\\", "/") + ".yaml"

        # 파일 열기 및 내용 수정
        with open(target_file, 'r') as file:
            lines = file.readlines()

        # 새로운 내용을 담을 리스트 생성
        new_lines = []
        for line in lines:
            # 'device:', 'test_vector_gen:', 'debug:' 부분을 수정
            if line.startswith('device:'):
                new_lines.append(f'device: {device}\n')
            elif line.startswith('test_vector_gen:'):
                new_lines.append(f'test_vector_gen: {test_vector_gen}\n')
            elif line.startswith('debug:'):
                new_lines.append(f'debug: {debug}\n')
            else:
                new_lines.append(line)

        # 수정된 내용을 다시 파일에 쓰기
        with open(target_file, 'w') as file:
            file.writelines(new_lines)

        PRINT_("파일 업데이트 완료")

    def stop(self):
        self._running = False
        self.quit()

        self.remove_all_container()
        self.wait(3000)


class FileCopyThread(QThread):
    progress_signal = pyqtSignal(int)  # 진행 상황 업데이트
    finished_signal = pyqtSignal(list)  # 완료 시 신호

    def __init__(self, test_model_paths, target_directory):
        super().__init__()
        self.test_model_paths = test_model_paths
        self.target_directory = target_directory
        self.all_test_paths = []

    def run(self):
        for cnt, test_model in enumerate(self.test_model_paths):
            directory, file_name = separate_folders_and_files(test_model)
            name, ext = separate_filename_and_extension(file_name)

            target_dir = os.path.join(self.target_directory, name).replace("\\", "/")
            CheckDir(target_dir)

            # 복사 작업
            src_file = test_model
            target_file = os.path.join(target_dir, file_name)
            shutil.copy2(src_file, target_file)
            self.all_test_paths.append(target_file)

            # 추가 파일 처리 (예: .prototxt)
            if "caffemodel" in ext:
                src_file = os.path.join(directory, f"{name}.prototxt")
                target_file = os.path.join(target_dir, f"{name}.prototxt")
                shutil.copy2(src_file, target_file)

            # 진행 상황 업데이트
            cnt = (cnt + 1) * 100 // len(self.test_model_paths)  # 퍼센트 계산
            self.progress_signal.emit(cnt)

        # 완료 신호
        self.finished_signal.emit(self.all_test_paths)


class Model_Verify_Class(QObject):
    send_sig_delete_all_sub_widget = pyqtSignal()
    send_sig_test = pyqtSignal()

    def __init__(self, parent, grand_parent):
        super().__init__()
        self.file_copy_thread = None
        self.parent = parent
        self.grand_parent = grand_parent

        self.result_thread = None
        self.start_evaluation_time = None
        self.end_evaluation_time = None
        self.work_progress = None
        self.user_error_fmt = None
        self.model_analyze_thread_instance = None
        self.added_scenario_widgets = None

        self.insert_widget_thread = None
        self.insert_widget_progress = None

        self.send_sig_delete_all_sub_widget.connect(self.update_all_sub_widget)
        self.send_sig_test.connect(self.testF)

    def open_file(self):
        # self.parent.mainFrame_ui.scenario_path_lineedit.setText(self.parent.directory.replace("\\", "/"))

        self.clear_sub_widget()

    def clear_sub_widget(self):
        while self.parent.mainFrame_ui.formLayout.count():
            item = self.parent.mainFrame_ui.formLayout.takeAt(0)
            widget = item.widget()
            if widget:
                widget.setParent(None)

        PRINT_("delete all")
        self.send_sig_delete_all_sub_widget.emit()

    def finish_insert_each_widget(self):
        if self.insert_widget_progress is not None:
            self.insert_widget_progress.close()
            print(f"total loaded models: {len(self.added_scenario_widgets)}")

    def insert_each_widget(self, cnt, target_model_path, repo_src, License="unknown"):
        if self.insert_widget_progress is not None:
            rt = load_module_func(module_name="source.ui_designer.main_widget")
            widget_ui = rt.Ui_Form()
            widget_instance = QtWidgets.QWidget()
            widget_ui.setupUi(widget_instance)

            model = os.path.basename(target_model_path)
            widget_ui.scenario_checkBox.setText(model)
            widget_ui.srclineEdit.setText(repo_src)
            widget_ui.licenselineEdit.setText(License)

            widget_ui.pathlineEdit.setText(f"{target_model_path}")
            widget_ui.parametersetting_textEdit.setMinimumHeight(200)
            self.parent.mainFrame_ui.formLayout.setWidget(cnt, QtWidgets.QFormLayout.FieldRole,
                                                          widget_instance)

            self.added_scenario_widgets.append((widget_ui, widget_instance))
            # self.insert_widget_progress.onCountChanged(value=cnt)

            widget_ui.inittextEdit.hide()
            widget_ui.conversiontextEdit.hide()
            widget_ui.compilertextEdit.hide()
            widget_ui.estimationtextEdit.hide()
            widget_ui.analysistextEdit.hide()
            widget_ui.profiletextEdit.hide()
            widget_ui.enntesttextEdit.hide()
            widget_ui.onnxruntime_InferenceSessiontextEdit.hide()

            # 체크박스 신호 연결
            widget_ui.logonoffcheckBox.stateChanged.connect(
                lambda state, ui=widget_ui: self.toggle_text_edits(ui, state)
            )

    @staticmethod
    def toggle_text_edits(ui, state):
        """체크박스 상태에 따라 위젯 숨기기/보이기"""
        if state == QtCore.Qt.Checked:
            ui.inittextEdit.show()
            ui.conversiontextEdit.show()
            ui.compilertextEdit.show()
            ui.estimationtextEdit.show()
            ui.analysistextEdit.show()
            ui.profiletextEdit.show()
            ui.enntesttextEdit.show()
            ui.onnxruntime_InferenceSessiontextEdit.show()
        else:
            ui.inittextEdit.hide()
            ui.conversiontextEdit.hide()
            ui.compilertextEdit.hide()
            ui.estimationtextEdit.hide()
            ui.analysistextEdit.hide()
            ui.profiletextEdit.hide()
            ui.enntesttextEdit.hide()
            ui.onnxruntime_InferenceSessiontextEdit.hide()

    def testF(self):
        # self.insert_widget_progress.setProgressBarMaximum(max_value=len(all_test_path))

        # BASE DIR 아래 Result 폴더 아래에 평가할 모델 복사
        user_fmt = [fmt.strip() for fmt in self.grand_parent.targetformat_lineedit.text().split(",") if fmt.strip()]

        # s = time.time()
        get_test_model_path = get_directory(base_dir=self.parent.directory, user_defined_fmt=user_fmt,
                                            file_full_path=True)
        # PRINT_("[get_directory] ===========================================>", time.time() - s)

        # Shared Volume 위치를 지정
        self.parent.directory = os.path.join(BASE_DIR, "Result").replace("\\", "/")
        self.parent.mainFrame_ui.scenario_path_lineedit.setText(self.parent.directory)
        CheckDir(self.parent.directory)

        all_test_path = []
        # s = time.time()
        PRINT_("+++++++++++++++++++++++++++++++++++++")
        for cnt, test_model in enumerate(get_test_model_path):
            directory, file_name = separate_folders_and_files(test_model)
            name, ext = separate_filename_and_extension(file_name)

            target_dir = os.path.join(self.parent.directory, name).replace("\\", "/")
            CheckDir(target_dir)

            src_file = test_model
            target_file = os.path.join(target_dir, file_name)
            shutil.copy2(src_file, target_file)

            all_test_path.append(target_file)

            if "caffemodel" in ext:
                src_file = os.path.join(directory, f"{name}.prototxt")
                target_file = os.path.join(target_dir, f"{name}.prototxt")
                shutil.copy2(src_file, target_file)

            cnt = cnt % 100
            PRINT_(cnt)
            self.insert_widget_progress.onCountChanged(value=cnt)

        # PRINT_("[move to Result] ===========================================>", time.time() - s)

        # if self.parent.mainFrame_ui.popctrl_radioButton.isChecked():
        #     self.insert_widget_progress = ProgressDialog(modal=False, message="Loading Scenario")
        # else:
        #     self.insert_widget_progress = ProgressDialog(modal=True, message="Loading Scenario")

        # self.insert_widget_progress.setProgressBarMaximum(max_value=len(all_test_path))

        repo_tag = self.parent.mainFrame_ui.dockerimagecomboBox.currentText()
        tag = int(repo_tag.split(":")[1].split(".")[0])
        if tag >= 7:
            full_file = os.path.join(BASE_DIR, "model_configuration", "Ver2.0_model_config_new.yaml")
        else:
            full_file = os.path.join(BASE_DIR, "model_configuration", "Ver1.0_model_config_new.yaml")

        with open(full_file, 'r') as model_config_file:
            model_config_data = self.parent.yaml.load(model_config_file)

        self.added_scenario_widgets = []
        self.insert_widget_thread = Load_Target_Dir_Thread(all_test_path, self.parent,
                                                           model_config_data=model_config_data)
        self.insert_widget_thread.send_scenario_update_ui_sig.connect(self.insert_each_widget)
        self.insert_widget_thread.send_finish_scenario_update_ui_sig.connect(self.finish_insert_each_widget)

        self.insert_widget_thread.start()

        # self.insert_widget_progress.show_progress()

    def on_file_copy_complete(self, all_test_path):
        # 복사 완료 후 후속 작업
        repo_tag = self.parent.mainFrame_ui.dockerimagecomboBox.currentText()
        tag = int(repo_tag.split(":")[1].split(".")[0])
        if tag >= 7:
            full_file = os.path.join(BASE_DIR, "model_configuration", "Ver2.0_model_config_new.yaml")
        else:
            full_file = os.path.join(BASE_DIR, "model_configuration", "Ver1.0_model_config_new.yaml")

        with open(full_file, 'r') as model_config_file:
            model_config_data = self.parent.yaml.load(model_config_file)

        self.added_scenario_widgets = []
        self.insert_widget_thread = Load_Target_Dir_Thread(all_test_path, self.parent,
                                                           model_config_data=model_config_data)
        self.insert_widget_thread.send_scenario_update_ui_sig.connect(self.insert_each_widget)
        self.insert_widget_thread.send_finish_scenario_update_ui_sig.connect(self.finish_insert_each_widget)
        self.insert_widget_thread.start()

    def update_all_sub_widget(self):
        PRINT_("send sig")

        # ProgressDialog 생성
        # if self.parent.mainFrame_ui.popctrl_radioButton.isChecked():
        #     self.insert_widget_progress = ProgressDialog(modal=False, message="Loading Scenario")
        # else:
        #     self.insert_widget_progress = ProgressDialog(modal=True, message="Loading Scenario")
        self.insert_widget_progress = ProgressDialog(modal=False, message="Loading Scenario")
        self.insert_widget_progress.show_progress()

        # BASE DIR 아래 Result 폴더 아래에 평가할 모델 복사
        user_fmt = [fmt.strip() for fmt in self.grand_parent.targetformat_lineedit.text().split(",") if fmt.strip()]
        get_test_model_path = get_directory(base_dir=self.parent.directory, user_defined_fmt=user_fmt,
                                            file_full_path=True)

        # Shared Volume 위치 지정
        self.parent.directory = os.path.join(BASE_DIR, "Result").replace("\\", "/")
        self.parent.mainFrame_ui.scenario_path_lineedit.setText(self.parent.directory)
        CheckDir(self.parent.directory)

        # 스레드 생성 및 실행
        self.file_copy_thread = FileCopyThread(get_test_model_path, self.parent.directory)
        self.file_copy_thread.progress_signal.connect(self.insert_widget_progress.onCountChanged)  # 진행 상황 업데이트
        self.file_copy_thread.finished_signal.connect(self.on_file_copy_complete)  # 완료 시 처리
        self.file_copy_thread.start()

    def set_text_progress(self, string):
        if self.work_progress is not None:
            self.work_progress.onProgressTextChanged(string)

    @staticmethod
    def update_timeout_result(execute_cmd, target_widget):
        if "init" in execute_cmd:
            target_widget[0].initlineEdit.setText("Runtime Out")
        if "conversion" in execute_cmd:
            target_widget[0].conversionlineEdit.setText("Runtime Out")
        elif "compile" in execute_cmd:
            target_widget[0].compilelineEdit.setText("Runtime Out")
        elif "estimation" in execute_cmd:
            target_widget[0].estimationlineEdit.setText("Runtime Out")
        elif "analysis" in execute_cmd:
            target_widget[0].analysislineEdit.setText("Runtime Out")
        elif "profiling" in execute_cmd:
            target_widget[0].profilinglineEdit.setText("Runtime Out")

    @staticmethod
    def update_onnx_info(sub_widget, onnx_info):
        sub_widget[0].modelvalidaty.setText(onnx_info["model check"][0])
        sub_widget[0].directshape.setText(onnx_info["direct shape"][0])
        sub_widget[0].simplifier.setText(onnx_info["simplifier"][0])
        sub_widget[0].onnxruntime_InferenceSession.setText(onnx_info["integrity"][0])
        sub_widget[0].onnxruntime_InferenceSessiontextEdit.setText(onnx_info["integrity"][1])
        sub_widget[0].modeltype.setText(onnx_info["precision"][0])
        sub_widget[0].onnxdomain.setText(onnx_info["domain"][0])
        sub_widget[0].onnxlineEdit.setText(onnx_info["opset_version"][0])

    @staticmethod
    def analyze_memory_trace(memory_usage):
        if len(memory_usage) == 0:
            return
        used_memory = ""
        # "Start"와 "End" 위치 찾기
        try:
            start_idx = memory_usage.index("Start")
            end_idx = memory_usage.index("End")

            # "Start"와 "End" 사이의 값 추출
            values_between = [x for x in memory_usage[start_idx + 1:end_idx] if isinstance(x, (int, float))]
            # 0 값을 제외한 값만 필터링
            filtered_values = [x for x in values_between if x != 0]

            # 평균, 최소, 최대 계산
            try:
                avg_val = sum(filtered_values) / len(filtered_values)
                min_val = min(filtered_values)
                max_val = max(filtered_values)
            except:
                # avg_val = 0
                # min_val = 0
                # max_val = 0
                PRINT_(f"division by zero")
                return used_memory

            # used_memory = f"Average: {avg_val:>10.1f}"
            used_memory = f"AVG.: {avg_val:>10.2f}\nMAX.: {max_val:>10.2f}\nMIN.: {min_val:>10.2f}"

        except ValueError as e:
            PRINT_(f"리스트에 'Start' 또는 'End'가 없습니다: {e}")

        return used_memory

    def update_test_result_2(self, sub_widget, execute_cmd, TestResult):
        if self.work_progress is not None:
            PRINT_(execute_cmd)
            result = TestResult[execute_cmd][0]
            log = TestResult[execute_cmd][1]

            memory_usage = ""
            if len(TestResult[execute_cmd][2]) != 0:
                memory_usage = self.analyze_memory_trace(memory_usage=TestResult[execute_cmd][2])

            if "init" in execute_cmd:
                sub_widget[0].initlineEdit.setText(result)
                sub_widget[0].parametersetting_textEdit.setText(log)

            elif "conversion" in execute_cmd:
                sub_widget[0].conversionlineEdit.setText(result)
                sub_widget[0].conversiontextEdit.setText(log)

            elif "compile" in execute_cmd:
                sub_widget[0].compilelineEdit.setText(result)
                sub_widget[0].compilertextEdit.setText(log)

            elif "estimation" in execute_cmd:
                sub_widget[0].estimationlineEdit.setText(result)
                sub_widget[0].estimationtextEdit.setText(log)

            elif "analysis" in execute_cmd:
                sub_widget[0].analysislineEdit.setText(result)
                sub_widget[0].analysistextEdit.setText(log)

            elif "profiling" in execute_cmd:
                sub_widget[0].profilinglineEdit.setText(result)
                sub_widget[0].profiletextEdit.setText(log)

            elif "enntest" in execute_cmd:
                # TextEdit와 LineEdit 초기화
                sub_widget[0].memorytextEdit.setText(memory_usage)

                sub_widget[0].enntestlineEdit.setText(result)
                if "success" in result.lower():
                    temp = log.split("\n")
                    for str_ in temp:
                        if "snr" in str_.lower():
                            try:
                                sub_widget[0].snrlineEdit.setText(
                                    re.sub(r"[\[\] ]", "", str_.split(":")[-1]).replace("dB", " dB"))
                            except:
                                sub_widget[0].snrlineEdit.setText("")
                        elif "perf" in str_.lower():
                            try:
                                sub_widget[0].exeperflineEdit.setText(
                                    re.sub(r"[\[\] ]", "", str_.split(":")[-1]).replace("fps", " fps"))
                            except:
                                sub_widget[0].exeperflineEdit.setText("")
                        elif "iter" in str_.lower() and "total" in str_.lower():
                            temp2 = str_.split("/")
                            try:
                                total_iter = re.sub(r"[\[\] ]", "", temp2[0].split(":")[-1])
                                sub_widget[0].iterlineEdit.setText(total_iter)

                                execution_time = round(
                                    float(re.sub(r"[\[\] ]", "", temp2[1].split(":")[-1]).replace("us", "")) / (
                                            int(total_iter) * 1000), 2)
                                sub_widget[0].exetimelineEdit.setText(str(execution_time) + " ms")

                                # sub_widget[0].iterlineEdit.setText(re.sub(r"[\[\] ]", "", temp2[0].split(":")[-1]))
                                # sub_widget[0].exetimelineEdit.setText(
                                #     re.sub(r"[\[\] ]", "", temp2[1].split(":")[-1]).replace("us", " us"))
                            except:
                                sub_widget[0].iterlineEdit.setText("")
                                sub_widget[0].exetimelineEdit.setText("")
                else:
                    sub_widget[0].enntesttextEdit.setText(log)

                # memory profile graph 그리기
                if len(TestResult[execute_cmd][2]) != 0:

                    # # 기존 그래프 제거 및 레이아웃 추가
                    # if not sub_widget[0].memoryprofilewidget.layout():
                    #     # 메모리 프로파일 위젯에 새로운 레이아웃 추가
                    #     layout = QtWidgets.QVBoxLayout(sub_widget[0].memoryprofilewidget)
                    #     sub_widget[0].memoryprofilewidget.setLayout(layout)
                    #
                    # # 기존 그래프 제거
                    # layout = sub_widget[0].memoryprofilewidget.layout()
                    # for i in reversed(range(layout.count())):
                    #     widget_to_remove = layout.itemAt(i).widget()
                    #     if widget_to_remove:
                    #         layout.removeWidget(widget_to_remove)  # 위젯을 레이아웃에서 제거
                    #         widget_to_remove.deleteLater()  # 메모리 해제
                    # 레이아웃 내 모든 위젯 삭제

                    # 새로운 그래프 생성
                    layout = sub_widget[0].memory_graph
                    figure = Figure()  # Matplotlib Figure 생성
                    canvas = FigureCanvas(figure)  # Figure를 Canvas로 변환
                    layout.addWidget(canvas)  # Canvas를 레이아웃에 추가

                    # 데이터 추출 및 처리
                    values_between = []
                    for val in TestResult[execute_cmd][2]:
                        if "Start" in str(val) or "End" in str(val):  # 특정 조건 건너뛰기
                            continue
                        try:
                            values_between.append(float(val))  # 숫자로 변환하여 리스트에 추가
                        except ValueError:
                            continue  # 숫자로 변환할 수 없는 값은 무시

                    # 그래프 그리기
                    ax = figure.add_subplot(111)  # Subplot 추가
                    if values_between:  # 데이터가 있을 경우에만 그래프 그리기
                        ax.plot(values_between, color='b', label="Memory Usage")  # 데이터 플롯팅
                        # ax.plot(values_between, marker='o', color='b', label="Memory Usage")  # 데이터 플롯팅
                        ax.set_title("Memory Usage Profile")  # 그래프 제목
                        ax.set_xlabel("Time")  # X축 라벨
                        ax.set_ylabel("Memory (MB)")  # Y축 라벨
                        ax.legend()  # 범례 추가
                        ax.grid()  # 격자 추가

                    # Canvas 업데이트
                    canvas.draw()

    def update_test_result(self, elapsed_time, sub_widget, executed_cnt, cwd):
        if self.work_progress is not None:
            self.work_progress.onCountChanged(value=executed_cnt)
            sub_widget[0].elapsedlineEdit.setText(elapsed_time)
            self.save(self_saving=True)

    def error_update_test_result(self, error_message, sub_widget):
        if self.work_progress is not None:
            self.work_progress.close()

        sub_widget[0].contexts_textEdit.setText(error_message)
        sub_widget[0].lineEdit.setText("Error")

    @staticmethod
    def find_and_stop_qthreads():
        app = QApplication.instance()
        if app:
            for widget in app.allWidgets():
                if isinstance(widget, QThread) and widget is not QThread.currentThread():
                    print(f"Stopping QThread: {widget}")
                    widget.quit()
                    widget.wait()

        # QObject 트리에서 QThread 찾기
        for obj in QObject.children(QApplication.instance()):
            if isinstance(obj, QThread) and obj is not QThread.currentThread():
                print(f"Stopping QThread: {obj}")
                obj.quit()
                obj.wait()

    @staticmethod
    def stop_all_threads():
        current_thread = threading.current_thread()

        for thread in threading.enumerate():
            if thread is current_thread:  # 현재 실행 중인 main 스레드는 제외
                continue

            if isinstance(thread, threading._DummyThread):  # 더미 스레드는 제외
                print(f"Skipping DummyThread: {thread.name}")
                continue

            print(f"Stopping Thread: {thread.name}")

            if hasattr(thread, "stop"):  # stop() 메서드가 있으면 호출
                thread.stop()
            elif hasattr(thread, "terminate"):  # terminate() 메서드가 있으면 호출
                thread.terminate()

            if thread.is_alive():
                thread.join(timeout=1)  # 1초 기다린 후 종료

    def finish_update_test_result(self, normal_stop):
        if self.work_progress is not None:
            self.work_progress.close()
            self.find_and_stop_qthreads()
            self.stop_all_threads()

            self.end_evaluation_time = time.time()
            elapsed_time = self.end_evaluation_time - self.start_evaluation_time
            days = elapsed_time // (24 * 3600)
            remaining_secs = elapsed_time % (24 * 3600)
            hours = remaining_secs // 3600
            remaining_secs %= 3600
            minutes = remaining_secs // 60
            seconds = remaining_secs % 60

            total_time = f"{int(days)}day {int(hours)}h {int(minutes)}m {int(seconds)}s"
            msg_box = QtWidgets.QMessageBox()

            if normal_stop:
                msg_box.setWindowTitle("Test Done...")
                msg_box.setText(f"All Test Done !\nSave Button to store result data\nElapsed time: {total_time}")
            else:
                msg_box.setWindowTitle("Stop Test...")
                msg_box.setText(f"User forcibly terminated !")

            msg_box.setStandardButtons(QtWidgets.QMessageBox.Yes)
            # Always show the message box on top
            if check_environment() == "Windows":
                msg_box.setWindowFlags(msg_box.windowFlags() | QtCore.Qt.WindowStaysOnTopHint)

            # 메시지 박스를 최상단에 표시
            answer = msg_box.exec_()

    def set_max_progress_cnt(self, max_cnt):
        if self.work_progress is not None:
            self.work_progress.setProgressBarMaximum(max_value=max_cnt)

    def stop_analyze(self):
        if self.work_progress is not None:
            self.work_progress.close()

        if self.model_analyze_thread_instance is not None:
            self.model_analyze_thread_instance.stop()

    def analyze(self):
        if self.added_scenario_widgets is None:
            return

        if len(self.added_scenario_widgets) == 0:
            return

        if check_environment() == "Windows":
            process_name = "EXCEL.EXE"
            try:
                subprocess.run(["taskkill", "/t", "/im", process_name, "/f"], check=True, stdout=subprocess.DEVNULL,
                               stderr=subprocess.DEVNULL)
            except subprocess.CalledProcessError:
                PRINT_(f"[-] Windows: '{process_name}' 프로세스를 찾을 수 없습니다.")
        else:
            process_name = "LibreOfficeCalc"
            try:
                subprocess.run(["killall", "-9", process_name], check=True, stdout=subprocess.DEVNULL,
                               stderr=subprocess.DEVNULL)
            except subprocess.CalledProcessError:
                PRINT_(f"[-] Linux: '{process_name}' 프로세스를 찾을 수 없습니다.")

        check = False
        for cnt, target_widget in enumerate(self.added_scenario_widgets):
            if target_widget[0].scenario_checkBox.isChecked():
                check = True

                target_widget[0].modelvalidaty.setText("")
                target_widget[0].onnxdomain.setText("")
                target_widget[0].onnxlineEdit.setText("")
                target_widget[0].modeltype.setText("")
                target_widget[0].onnxruntime_InferenceSession.setText("")
                target_widget[0].onnxruntime_InferenceSessiontextEdit.setText("")
                target_widget[0].directshape.setText("")
                target_widget[0].simplifier.setText("")

                # success/ fail lineedit 초기화
                target_widget[0].initlineEdit.setText("")
                target_widget[0].inittextEdit.setText("")

                target_widget[0].conversionlineEdit.setText("")
                target_widget[0].conversiontextEdit.setText("")

                target_widget[0].compilelineEdit.setText("")
                target_widget[0].compilertextEdit.setText("")

                target_widget[0].estimationlineEdit.setText("")
                target_widget[0].estimationtextEdit.setText("")

                target_widget[0].analysislineEdit.setText("")
                target_widget[0].analysistextEdit.setText("")

                target_widget[0].profilinglineEdit.setText("")
                target_widget[0].profiletextEdit.setText("")

                target_widget[0].elapsedlineEdit.setText("")
                target_widget[0].parametersetting_textEdit.setText("")

                target_widget[0].enntestlineEdit.setText("")
                target_widget[0].enntesttextEdit.setText("")

                target_widget[0].memorytextEdit.setText("")
                target_widget[0].iterlineEdit.setText("")
                target_widget[0].snrlineEdit.setText("")
                target_widget[0].exetimelineEdit.setText("")
                target_widget[0].exeperflineEdit.setText("")

                # 레이아웃 내 모든 위젯 삭제
                layout = target_widget[0].memory_graph
                for i in reversed(range(layout.count())):
                    widget = layout.itemAt(i).widget()  # 레이아웃에서 위젯 가져오기
                    if widget is not None:
                        layout.removeWidget(widget)  # 레이아웃에서 위젯 제거
                        widget.setParent(None)  # 부모 위젯 참조 제거하여 삭제 준비
                        del widget  # 위젯 삭제
                QApplication.processEvents()  # GUI를 즉시 갱신하여 삭제된 위젯 반영

                # # mrmory profile 레이아웃이 존재하면 기존 그래프 및 위젯 제거
                # layout = target_widget[0].memoryprofilewidget.layout()
                # if layout:
                #     # 레이아웃에서 위젯을 하나씩 제거
                #     for i in reversed(range(layout.count())):
                #         widget_to_remove = layout.itemAt(i).widget()
                #         if widget_to_remove:
                #             layout.removeWidget(widget_to_remove)  # 레이아웃에서 위젯 제거
                #             widget_to_remove.deleteLater()  # 메모리에서 해제

                #     # 레이아웃을 새롭게 다시 설정
                #     new_layout = QtWidgets.QVBoxLayout(target_widget[0].memoryprofilewidget)
                #     target_widget[0].memoryprofilewidget.setLayout(new_layout)  # 새 레이아웃 설정
                # else:
                #     # 레이아웃이 없으면 새 레이아웃 추가
                #     layout = QtWidgets.QVBoxLayout(target_widget[0].memoryprofilewidget)
                #     target_widget[0].memoryprofilewidget.setLayout(layout)  # 새 레이아웃 설정

                directory, model_name = separate_folders_and_files(target_widget[0].pathlineEdit.text())
                filename, extension = separate_filename_and_extension(model_name)

                remove_alldata_files_except_specific_extension(directory=directory,
                                                               extension=extension.replace(".", ""))

        if not check:
            msg_box = QtWidgets.QMessageBox()  # QMessageBox 객체 생성
            msg_box.setWindowTitle("Check Test Model")  # 대화 상자 제목
            msg_box.setText(
                "test model is required.\nMark model")
            msg_box.setStandardButtons(QtWidgets.QMessageBox.Yes)  # Yes/No 버튼 추가

            if check_environment() == "Windows":
                msg_box.setWindowFlags(msg_box.windowFlags() | QtCore.Qt.WindowStaysOnTopHint)  # 항상 위에 표시

            answer = msg_box.exec_()  # 대화 상자를 실행하고 사용자의 응답을 반환
            if answer == QtWidgets.QMessageBox.Yes:
                return

        self.start_evaluation_time = time.time()

        if self.parent.mainFrame_ui.popctrl_radioButton.isChecked():
            self.work_progress = ProgressDialog(modal=False, message="Analyzing AI Model", show=True)
        else:
            self.work_progress = ProgressDialog(modal=True, message="Analyzing AI Model", show=True)

        self.work_progress.send_user_close_event.connect(self.stop_analyze)

        self.user_error_fmt = [fmt.strip() for fmt in self.grand_parent.error_lineedit.text().split(",")]

        repo_tag = self.parent.mainFrame_ui.dockerimagecomboBox.currentText()
        self.model_analyze_thread_instance = Model_Analyze_Thread(self, self.grand_parent, repo_tag=repo_tag)
        self.model_analyze_thread_instance.output_signal.connect(self.update_test_result)
        self.model_analyze_thread_instance.output_signal_2.connect(self.update_test_result_2)
        self.model_analyze_thread_instance.send_onnx_opset_ver_signal.connect(self.update_onnx_info)
        self.model_analyze_thread_instance.timeout_output_signal.connect(self.update_timeout_result)

        self.model_analyze_thread_instance.send_set_text_signal.connect(self.set_text_progress)
        self.model_analyze_thread_instance.send_max_progress_cnt.connect(self.set_max_progress_cnt)
        self.model_analyze_thread_instance.finish_output_signal.connect(self.finish_update_test_result)

        self.model_analyze_thread_instance.start()

        self.work_progress.show_progress()

    def select_all_scenario(self, check):
        if self.added_scenario_widgets is None or len(self.added_scenario_widgets) == 0:
            return

        for scenario_widget, scenario_widget_instance in self.added_scenario_widgets:
            scenario_widget.scenario_checkBox.setChecked(check)

    def save(self, self_saving=False):
        if self.added_scenario_widgets is None or len(self.added_scenario_widgets) == 0:
            return

        def clean_data(value):
            if isinstance(value, str):
                # Excel에서 허용되지 않는 문자 제거 (\x00 같은 제어 문자)
                # value = re.sub(r"[\x00-\x1F\x7F-\x9F]", "", value)
                # \n과 \t는 제외하고 제어 문자 제거
                value = re.sub(r"[^\n\t\x20-\x7E]", "", value)
            return value

        result = []

        for cnt, target_widget in enumerate(self.added_scenario_widgets):
            # ret = target_widget[0].scenario_checkBox.text().strip().split(".")
            model, framework = os.path.splitext(target_widget[0].scenario_checkBox.text())
            #
            widget_result = {
                "Model": clean_data(model),
                "Framework": clean_data(framework.replace(".", "")),
                "License": clean_data(target_widget[0].licenselineEdit.text().strip()),
                "Set parameter": clean_data(target_widget[0].parametersetting_textEdit.toPlainText().strip()),
                "Check Model": clean_data(target_widget[0].modelvalidaty.text().strip()),
                "Direct Shape": clean_data(target_widget[0].directshape.text().strip()),
                "Simplifier": clean_data(target_widget[0].simplifier.text().strip()),
                "Onnx_Model_RuntimeInferenceSession_Check": clean_data(
                    target_widget[0].onnxruntime_InferenceSession.text().strip()),
                "Onnx_Model_RuntimeInferenceSession_Check_log": clean_data(
                    target_widget[0].onnxruntime_InferenceSessiontextEdit.toPlainText()),
                "Onnx_Model_Precision": clean_data(target_widget[0].modeltype.text().strip()),
                "Onnx_Domain": clean_data(target_widget[0].onnxdomain.text().strip()),
                "Onnx_Opset Version": clean_data(target_widget[0].onnxlineEdit.text().strip()),

                "init_Result": clean_data(target_widget[0].initlineEdit.text().strip()),
                "init_log": clean_data(""),  # 초기화된 값이 비어 있다면
                "conversion_Result": clean_data(target_widget[0].conversionlineEdit.text().strip()),
                "conversion_log": clean_data(target_widget[0].conversiontextEdit.toPlainText()),
                "compile_Result": clean_data(target_widget[0].compilelineEdit.text().strip()),
                "compile_log": clean_data(target_widget[0].compilertextEdit.toPlainText()),
                "estimation_Result": clean_data(target_widget[0].estimationlineEdit.text().strip()),
                "estimation_log": clean_data(target_widget[0].estimationtextEdit.toPlainText()),
                "analysis_Result": clean_data(target_widget[0].analysislineEdit.text().strip()),
                "analysis_log": clean_data(target_widget[0].analysistextEdit.toPlainText()),
                "profiling_Result": clean_data(target_widget[0].profilinglineEdit.text().strip()),
                "profiling_log": clean_data(target_widget[0].profiletextEdit.toPlainText()),
                "enntest_Result": clean_data(target_widget[0].enntestlineEdit.text().strip()),
                "enntest_log": clean_data(target_widget[0].enntesttextEdit.toPlainText()),
                "enntest_Iter.[count]": clean_data(target_widget[0].iterlineEdit.text().strip()),
                "Memory Usage[MB]": clean_data(target_widget[0].memorytextEdit.toPlainText()),
                "enntest_SNR[dB]": clean_data(target_widget[0].snrlineEdit.text().strip()),
                # "enntest Total Execution Time[us]": clean_data(target_widget[0].exetimelineEdit.text().strip()),
                "enntest Execution Time[ms]": clean_data(target_widget[0].exetimelineEdit.text().strip()),
                "enntest Execution Performance[fps]": clean_data(target_widget[0].exeperflineEdit.text().strip()),
                "model_source": clean_data(target_widget[0].srclineEdit.text().strip()),
                "elapsed_time": clean_data(target_widget[0].elapsedlineEdit.text().strip()),
            }
            result.append(widget_result)

        result_path = os.path.join(BASE_DIR, "Result", "result.json")
        json_dump_f(result_path, result)

        df = pd.DataFrame(result)

        # df = pd.DataFrame(result).applymap(clean_data)
        # 엑셀 파일로 저장
        excel_file = result_path.replace("json", "xlsx")
        df.to_excel(excel_file, index=False, engine='openpyxl', startrow=1)
        # Load the workbook and access the active sheet
        wb = load_workbook(excel_file)
        ws = wb.active

        # 첫 번째 셀(A1)에 "Hello" 추가
        repo_tag = self.parent.mainFrame_ui.dockerimagecomboBox.currentText()
        ws.cell(row=1, column=1, value=str(repo_tag))

        # 줄 바꿈 설정 및 열 너비 자동 조정
        for row in ws.iter_rows(min_row=2):  # 데이터는 두 번째 줄부터 시작
            for cell in row:
                if cell.value:
                    cell.value = str(cell.value).replace("\\n", "\n")
                    cell.alignment = Alignment(wrap_text=True, vertical='top')

        # Auto-adjust column width
        for col in range(1, len(df.columns) + 1):
            max_length = 0
            column = get_column_letter(col)

            for row in ws.iter_rows(min_col=col, max_col=col):
                for cell in row:
                    if cell.value:  # 셀 값이 존재할 경우에만 길이 계산
                        max_length = max(max_length, len(str(cell.value)))

            ws.column_dimensions[column].width = max_length + 5  # Add padding to prevent truncation

        # Auto-adjust row height based on content (to ensure proper wrap)
        for row in ws.iter_rows(min_row=2):
            max_row_height = 0
            for cell in row:
                if cell.alignment.wrap_text and cell.value:  # 줄 바꿈이 설정된 셀에 대해 높이 조정
                    content_height = len(str(cell.value).split("\n"))
                    max_row_height = max(max_row_height, content_height)
            ws.row_dimensions[row[0].row].height = max_row_height * 15  # 행 높이 조정

        # Save the workbook with adjustments
        wb.save(excel_file)

        if not self_saving:
            msg_box = QtWidgets.QMessageBox()  # QMessageBox 객체 생성
            msg_box.setWindowTitle("Save Result")  # 대화 상자 제목
            msg_box.setText(
                "All saved.               \n")
            msg_box.setStandardButtons(QtWidgets.QMessageBox.Yes)  # Yes/No 버튼 추가

            if check_environment() == "Windows":
                msg_box.setWindowFlags(msg_box.windowFlags() | QtCore.Qt.WindowStaysOnTopHint)  # 항상 위에 표시

            answer = msg_box.exec_()  # 대화 상자를 실행하고 사용자의 응답을 반환


class Project_MainWindow(QtWidgets.QMainWindow):

    def __init__(self):
        super().__init__()

        self.dialog = None
        self.directory = None
        self.single_op_ctrl = None
        self.Shared_Volume = None
        self.repo_tag = None
        self.load_progress = None
        self.load_thread = None
        """ for main frame & widget """
        self.mainFrame_ui = None
        self.widget_ui = None

        self.setupUi()

        self.yaml = YAML()
        self.yaml.preserve_quotes = True  # 기존 따옴표 보존
        self.yaml.default_flow_style = None  # 블록 스타일로 저장

    def setupUi(self):
        # Load the main window's UI module
        rt = load_module_func(module_name="source.ui_designer.main_frame")
        self.mainFrame_ui = rt.Ui_MainWindow()
        self.mainFrame_ui.setupUi(self)

        element_list = [fmt.strip() for fmt in keyword["test_model"]]
        text_concatenation = ', '.join(element_list)
        self.mainFrame_ui.targetformat_lineedit.setText(text_concatenation)

        element_list = [fmt.strip() for fmt in keyword["error_keyword"]]
        text_concatenation = ', '.join(element_list)
        self.mainFrame_ui.error_lineedit.setText(text_concatenation)

        # element_list = [fmt.strip() for fmt in keyword["op_exe_cmd"]]
        # text_concatenation = ', '.join(element_list)
        # self.mainFrame_ui.command_lineedit.setText(text_concatenation)

        device_m_path = os.path.join(BASE_DIR, "model_configuration", "device_manager.json").replace("\\", "/")
        _, deviceID_data = json_load_f(file_path=device_m_path)
        self.mainFrame_ui.localdeviceidlineEdit.setText(deviceID_data["local device"])
        self.mainFrame_ui.sshdevicelineEdit.setText(deviceID_data["ssh device"])

        # image history
        history = os.path.join(BASE_DIR, "source", "history", "release_history.json")
        _, history_data = json_load_f(file_path=history)
        for cnt, data_ in enumerate(history_data):
            for item_num, (key, value) in enumerate(data_.items()):
                label = QtWidgets.QLabel(str(value))  # QLabel에 텍스트 추가
                font = label.font()
                font.setBold(False)  # 폰트를 기본 스타일로 설정 (굵기 제거)
                label.setFont(font)
                label.setAlignment(QtCore.Qt.AlignCenter)  # 가운데 정렬
                self.mainFrame_ui.history_table.addWidget(label, cnt + 1, item_num)

        self.mainFrame_ui.cmdlabel.hide()
        self.mainFrame_ui.command_lineedit.hide()

        self.setWindowTitle(Version)

    @staticmethod
    def kill_docker_desktop():
        L_processor = ["Docker Desktop.exe"]

        for proc in L_processor:
            try:
                # taskkill 명령어 실행
                result = subprocess.run(
                    ["taskkill", "/F", "/IM", rf"{proc}"],
                    capture_output=True,
                    text=True
                )
                if result.returncode == 0:
                    PRINT_("'Docker Desktop.exe' has been terminated successfully.")
                else:
                    PRINT_(f"Failed to terminate 'Docker Desktop.exe': {result.stderr}")
            except Exception as e:
                PRINT_(f"An error occurred: {e}")

    def closeEvent(self, event):
        answer = QtWidgets.QMessageBox.question(self,
                                                "Confirm Exit...",
                                                "Are you sure you want to exit?\nAll data will be lost.",
                                                QtWidgets.QMessageBox.Yes | QtWidgets.QMessageBox.No)

        if answer == QtWidgets.QMessageBox.Yes:
            # self.kill_docker_desktop()
            event.accept()
        else:
            event.ignore()

    def normalOutputWritten(self, text):
        cursor = self.mainFrame_ui.logtextbrowser.textCursor()
        cursor.movePosition(QtGui.QTextCursor.End)

        # 기본 글자 색상 설정
        color_format = cursor.charFormat()

        keywords = ['Error Code:', 'Error code:', 'Error msg:']
        color_format.setForeground(
            QtCore.Qt.red if any(keyword in text for keyword in keywords) else QtCore.Qt.black)

        cursor.setCharFormat(color_format)
        cursor.insertText(text)

        # 커서를 최신 위치로 업데이트
        self.mainFrame_ui.logtextbrowser.setTextCursor(cursor)
        self.mainFrame_ui.logtextbrowser.ensureCursorVisible()

    def cleanLogBrowser(self):
        self.mainFrame_ui.logtextbrowser.clear()

    def connectSlotSignal(self):
        """ sys.stdout redirection """
        sys.stdout = EmittingStream(textWritten=self.normalOutputWritten)
        self.mainFrame_ui.log_clear_pushButton.clicked.connect(self.cleanLogBrowser)

        # evaluation tab
        self.mainFrame_ui.open_pushButton.clicked.connect(self.open_directory)
        self.mainFrame_ui.analyze_pushButton.clicked.connect(self.start_analyze)
        self.mainFrame_ui.all_check_scenario.clicked.connect(self.check_all_scenario)
        self.mainFrame_ui.all_uncheck_scenario.clicked.connect(self.check_all_scenario)
        self.mainFrame_ui.save_pushButton.clicked.connect(self.save_result)

        self.mainFrame_ui.actionOn.triggered.connect(self.log_browser_ctrl)
        self.mainFrame_ui.actionOff.triggered.connect(self.log_browser_ctrl)
        self.mainFrame_ui.actionOpen_Result_Excel.triggered.connect(self.open_test_result)

        self.mainFrame_ui.imageregpushButton.clicked.connect(self.image_registration)

        self.mainFrame_ui.configpushButton.clicked.connect(self.open_model_config)

        self.mainFrame_ui.remoteradioButton.toggled.connect(self.on_radio_button_toggled)
        self.mainFrame_ui.localradioButton.toggled.connect(self.on_radio_button_toggled)
        self.mainFrame_ui.cmd5.setEnabled(False)

        self.mainFrame_ui.localdeviceidpushButton.clicked.connect(self.save_deviceID)
        self.mainFrame_ui.sshdeviceidpushButton.clicked.connect(self.save_deviceID)

        self.single_op_ctrl = Model_Verify_Class(parent=self, grand_parent=self.mainFrame_ui)

        self.mainFrame_ui.groupBox_4.hide()

    def save_deviceID(self):
        device_m_path = os.path.join(BASE_DIR, "model_configuration", "device_manager.json").replace("\\", "/")

        dump_data = {
            "local device": f"{self.mainFrame_ui.localdeviceidlineEdit.text().strip()}",
            "ssh device": f"{self.mainFrame_ui.sshdevicelineEdit.text().strip()}"
        }

        with open(device_m_path, 'w', encoding='utf-8') as file:
            json.dump(dump_data, file, indent=4, ensure_ascii=False)

    def on_radio_button_toggled(self):
        env = check_environment()
        if env == "Windows":
            return

        if self.mainFrame_ui.remoteradioButton.isChecked():
            self.mainFrame_ui.cmd5.setChecked(False)
            self.mainFrame_ui.cmd5.setEnabled(False)
        else:
            self.mainFrame_ui.cmd5.setEnabled(True)

    def save_changes(self, full_file, content, dialog):
        # YAML 문자열을 파싱하여 데이터로 변환
        try:
            # StringIO를 사용해 문자열을 YAML 객체로 로드
            yaml_data = self.yaml.load(StringIO(content))

            # 수정된 내용을 파일에 저장
            with open(full_file, 'w') as file:
                self.yaml.dump(yaml_data, file)

            # 성공 메시지 (필요시)
            PRINT_("파일이 성공적으로 저장되었습니다.")
            dialog.accept()  # 저장 후 다이얼로그 닫기

        except Exception as e:
            # 에러 메시지 (필요시)
            PRINT_(f"저장 중 오류 발생: {e}")

    def open_model_config(self):
        repo_tag = self.mainFrame_ui.dockerimagecomboBox.currentText()
        tag = int(repo_tag.split(":")[1].split(".")[0])
        if tag >= 7:
            full_file = os.path.join(BASE_DIR, "model_configuration", "Ver2.0_model_config_new.yaml")
        else:
            full_file = os.path.join(BASE_DIR, "model_configuration", "Ver1.0_model_config_new.yaml")

        with open(full_file, 'r') as model_config_file:
            model_config_data = self.yaml.load(model_config_file)

        # YAML 데이터를 문자열로 변환
        model_config_str = StringIO()
        self.yaml.dump(model_config_data, model_config_str)
        model_config_str = model_config_str.getvalue()  # 문자열로 변환

        rt = load_module_func(module_name="source.ui_designer.fileedit")

        self.dialog = QtWidgets.QDialog()
        ui = rt.Ui_Dialog()
        ui.setupUi(self.dialog)
        # 시그널 슬롯 연결 람다 사용해서 직접 인자를 넘기자...........
        ui.fileedit_save.clicked.connect(lambda: self.save_changes(full_file, ui.textEdit.toPlainText(), self.dialog))
        ui.fileedit_cancel.clicked.connect(self.dialog.close)

        # QSyntaxHighlighter를 통해 ':'로 끝나는 줄을 파란색으로 강조 표시
        highlighter = ColonLineHighlighter(ui.textEdit.document())

        ui.textEdit.setPlainText(model_config_str)  # 변환된 문자열 설정
        self.dialog.setWindowModality(QtCore.Qt.ApplicationModal)
        self.dialog.show()

    @staticmethod
    def is_image_loaded():
        # docker image ls 명령어로 로드된 이미지 목록 가져오기
        list_cmd = 'docker image ls --format "{{.Repository}}:{{.Tag}}"'
        loaded_images, _, _ = user_subprocess(cmd=list_cmd, run_time=False, log=False)

        return loaded_images

    def image_registration(self):
        # load_file = easygui.fileopenbox(
        #     msg="Select AI Studio Image",
        #     title="Image Load",
        #     default="*.image",
        #     filetypes=["*.image", "*.tar"]
        # )
        load_file, _ = QFileDialog.getOpenFileName(
            None,  # 부모 위젯 없음
            "Image Load",  # 다이얼로그 제목
            "",  # 초기 디렉토리 (""은 현재 디렉토리)
            "AI Studio Image (*.image);;Tar Files (*.tar);;All Files (*)"  # 파일 필터
        )

        if load_file is None:
            self.update_docker_imageList()
            return

        repo_tag = get_image_info_from_dockerImg(load_file.replace("\\", "/"))
        repo_tag_in_drive = self.is_image_loaded()

        if not repo_tag or not repo_tag_in_drive:
            result = False
        else:
            result = any(element_a in element_b for element_a in repo_tag for element_b in repo_tag_in_drive)

        if result:
            PRINT_(f"Image '{repo_tag}' is already loaded. Skipping load.")
            self.imageLoadResult(success=True, elapsed_time=1)
            return

        message = f"Want to load '{repo_tag}'?."
        yes_ = Open_QMessageBox(message=message)

        if yes_:
            if self.mainFrame_ui.popctrl_radioButton.isChecked():
                self.load_progress = ProgressDialog(modal=False, message="Image Loading", parent=self)
            else:
                self.load_progress = ProgressDialog(modal=True, message="Image Loading", parent=self)

            # 드라이브 문자 추출 및 변환
            drive, path = os.path.splitdrive(load_file)
            mnt_path = "/mnt/" + drive.lower().replace(":", "") + path.replace("\\", "/")

            load_command = f"docker load -i {mnt_path}"
            self.load_thread = ImageLoadThread(parent=self, arg1=load_command, arg2=True)
            self.load_thread.send_finish_ui_sig.connect(self.imageLoadResult)
            self.load_thread.start()

            self.load_progress.show_progress()

    def update_docker_imageList(self):
        self.mainFrame_ui.dockerimagecomboBox.clear()
        self.mainFrame_ui.dockerimagecomboBox.addItems(self.is_image_loaded())

    def imageLoadResult(self, success, elapsed_time):
        if self.load_progress is not None:
            self.load_progress.close()

        if success:
            message = f"[Info] 이미지가 성공적으로 로드되었습니다.\nelapsed_time: {elapsed_time}"
        else:
            message = "[Warning] 이미지 로딩 실패 입니다."

        yes_ = Open_QMessageBox(message=message, no_b=False)

        self.update_docker_imageList()
        # self.update_containerList()

    def select_docker_image(self):
        # load_file = easygui.fileopenbox(
        #     msg="Select Test Set Scenario",
        #     title="Test Set Selection",
        #     default="*.image",
        #     filetypes=["*.image"]
        # )
        # QFileDialog for Single File Selection
        load_file, _ = QFileDialog.getOpenFileName(
            None,  # 부모 위젯 없음
            "Test Set Selection",  # 다이얼로그 제목
            "",  # 초기 디렉토리 (""은 현재 디렉토리)
            "Image Files (*.image);;All Files (*)"  # 파일 필터
        )

        if load_file is None:
            return

        image_name = self.mainFrame_ui.replineEdit.text().strip()
        image_tag = self.mainFrame_ui.taglineEdit.text().strip()

        check_command = f"docker image ls --format '{{{{.Repository}}}}:{{{{.Tag}}}}'"
        out, error, _ = user_subprocess(cmd=check_command, run_time=False)

        if f"{image_name}:{image_tag}" in out:
            PRINT_(f"{image_name}:{image_tag} 이미 로드되어 있습니다.")
        else:
            message = "이미지가 로드 되어 있지 않습니다. 로드 하시겠습니까?\n로딩하는데 최소 몇 분 이상이 소요가 됩니다"
            yes_ = Open_QMessageBox(message=message)

            if yes_:
                load_command = f"docker load -i {load_file}"
                load_result, _, _ = user_subprocess(cmd=load_command, run_time=False)
                if any("Loaded image:" in line for line in load_result):
                    PRINT_(f"{image_name}:{image_tag} 이미지가 성공적으로 로드되었습니다.")
                    self.mainFrame_ui.aistudiolineEdit.setText(os.path.basename(load_file))
                else:
                    PRINT_(f"{image_name}:{image_tag} 로딩 Fail 입니다.")

    def log_browser_ctrl(self):
        sender = self.sender()
        if sender:
            if sender.objectName() == "actionOff":
                self.mainFrame_ui.logtextbrowser.hide()
            else:
                self.mainFrame_ui.logtextbrowser.show()

    @staticmethod
    def open_excel():
        env = check_environment()
        if env == "Windows":
            excel_path = rf"C:\Program Files\Microsoft Office\root\Office16\EXCEL.EXE".replace("\\", "/")
        else:
            excel_path = "libreoffice"

        result_path = os.path.join(BASE_DIR, "Result", "result.xlsx").replace("\\", "/")

        if os.path.exists(result_path):
            subprocess.Popen([excel_path, result_path], shell=False)

    def open_test_result(self):
        thread = threading.Thread(target=self.open_excel, daemon=True)
        thread.start()

    # @staticmethod
    # def Xopen_test_result():
    #     env = check_environment()
    #     if env == "Windows":
    #         excel_path = rf"C:\Program Files\Microsoft Office\root\Office16\EXCEL.EXE".replace("\\", "/")
    #     else:
    #         excel_path = "libreoffice"
    #
    #     result_path = os.path.join(BASE_DIR, "Result", "result.xlsx").replace("\\", "/")
    #     if os.path.exists(result_path):
    #         subprocess.run([excel_path, result_path])

    def open_directory(self):
        # _directory = easygui.diropenbox()
        # QFileDialog for Directory Selection
        _directory = QFileDialog.getExistingDirectory(self, "Select Directory")
        if _directory:
            PRINT_("Selected Directory:", _directory)

        if _directory is None:
            return

        self.directory = _directory.replace("\\", "/")
        if self.single_op_ctrl is not None:
            self.single_op_ctrl.open_file()

    def start_analyze(self):
        if self.single_op_ctrl is not None:
            self.single_op_ctrl.analyze()

    def check_all_scenario(self):
        sender = self.sender()
        check = False

        if sender:
            if sender.objectName() == "all_check_scenario":
                check = True
            elif sender.objectName() == "all_uncheck_scenario":
                check = False

        self.single_op_ctrl.select_all_scenario(check=check)

    def save_result(self):
        if self.single_op_ctrl is not None:
            self.single_op_ctrl.save()


if __name__ == "__main__":
    env = check_environment()
    if env == "Windows":
        start_docker_desktop()

    import sys

    app = QtWidgets.QApplication(sys.argv)  # QApplication 생성 (필수)

    app.setStyle("Fusion")
    ui = Project_MainWindow()
    ui.showMaximized()
    ui.connectSlotSignal()
    ui.update_docker_imageList()

    sys.exit(app.exec_())
